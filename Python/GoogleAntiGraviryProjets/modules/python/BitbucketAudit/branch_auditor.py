import json
import requests
import os
import sys
from datetime import datetime
from requests.auth import HTTPBasicAuth

# Import common utilities
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from common.config_loader import load_config

def check_branch_delta(base_url, user, token, project, repo, source, target):
    """
    Checks if there are commits in Source NOT in Target.
    API: GET /rest/api/1.0/projects/{projectKey}/repos/{repositorySlug}/compare/commits
    params: from=source, to=target
    """
    url = f"{base_url}/rest/api/1.0/projects/{project}/repos/{repo}/commits"
    # Note: Bitbucket Server API for 'compare' might vary by version. 
    # Standard endpoint often used: /rest/api/1.0/projects/{project}/repos/{repo}/compare/commits?from={source}&to={target}
    # However, 'compare/commits' shows commits reachable from 'from' but not 'to'. 
    # Usually: 'from' is the feature branch, 'to' is master.
    # We want: Commits in Main (Source) NOT in CR (Target).
    # So we want distinct commits reachable from Source, excluding Target.
    # Param 'from' = Source, 'to' = Target.
    
    compare_url = f"{base_url}/rest/api/1.0/projects/{project}/repos/{repo}/compare/commits"
    params = {
        'from': source,
        'to': target,
        'limit': 100
    }
    
    try:
        response = requests.get(compare_url, auth=HTTPBasicAuth(user, token), params=params, timeout=10)
        if response.status_code == 200:
            data = response.json()
            # If 'values' has items, it means there are commits in Source not in Target
            return data.get('values', [])
        else:
            print(f"[-] API Error {response.status_code} for {repo}: {response.text}")
            return None
    except Exception as e:
        print(f"[-] Connection Error: {e}")
        return None

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_report(audit_results, base_url=""):
    """
    Generates report in both HTML and Excel format.
    Includes Links to Bitbucket Commits.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timestamp_file = datetime.now().strftime("%Y%m%d_%H%M")
    
    # -------------------------------------------------------------
    # 1. HTML REPORT
    # -------------------------------------------------------------
    html = f"""
    <html>
    <head>
        <title>Bitbucket Branch Audit Report</title>
        <style>
            body {{ font-family: Arial, sans-serif; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; }}
            .risk {{ color: red; font-weight: bold; }}
            .safe {{ color: green; }}
            a {{ text-decoration: none; color: #0052cc; }}
            a:hover {{ text-decoration: underline; }}
        </style>
    </head>
    <body>
        <h1>Bitbucket Branch Audit Report</h1>
        <p>Generated at: {timestamp}</p>
        <p>Purpose: Identify missing merges from Main Release Branch to CR Branches (Production Risk Check).</p>
        
        <table>
            <tr>
                <th>Country/Project</th>
                <th>Repo</th>
                <th>Source (Main)</th>
                <th>Target (CR)</th>
                <th>Status</th>
                <th>Missing Commits (Delta)</th>
            </tr>
    """
    
    # Prepare data for Excel
    excel_rows = []
    
    for res in audit_results:
        country = res['country']
        project = res['project']
        repo = res['repo']
        source = res['source']
        target = res['target']
        deltas = res['deltas']
        
        # Build Bitbucket Commit URL base: {base_url}/projects/{project}/repos/{repo}/commits/{commitId}
        repo_url_base = f"{base_url}/projects/{project}/repos/{repo}/commits"
        
        if deltas:
            status_html = f"<span class='risk'>MISSING {len(deltas)} COMMITS</span>"
            status_text = f"MISSING {len(deltas)} COMMITS"
            
            details_html = "<ul>"
            details_excel = ""
            
            for i, c in enumerate(deltas):
                commit_id = c['displayId']
                full_hash = c.get('id', commit_id) # API usually returns 'id' as full hash
                msg = c['message'].split('\n')[0] # Get first line only
                author = c['author']['name']
                link = f"{repo_url_base}/{full_hash}"
                
                # HTML List Item
                details_html += f"<li><a href='{link}' target='_blank'>{commit_id}</a>: {msg} ({author})</li>"
                
                # Excel List Item (just text for now, maybe links later)
                details_excel += f"{commit_id}: {msg} ({author})\n"
                
                # Add individual row to Excel if we want detailed view? 
                # Or keep summary view. Instructions imply "column with Link... so easier to have 1 row per missing commit?"
                # User asked "add one more column with Link to Bitbucket...". 
                # If there are many commits, creating 1 row per commit is clearer for Excel filtering.
                excel_rows.append({
                    "Timestamp": timestamp,
                    "Country": country,
                    "Project": project,
                    "Repository": repo,
                    "Source Branch": source,
                    "Target Branch": target,
                    "Status": "MISSING",
                    "Commit ID": commit_id,
                    "Description": msg,
                    "Author": author,
                    "Bitbucket Link": link
                })

            details_html += "</ul>"
            
        else:
            status_html = "<span class='safe'>SYNCED</span>"
            status_text = "SYNCED"
            details_html = "None. Target includes all Source changes."
            
            # For Synced, just add one row to show it was checked
            excel_rows.append({
                "Timestamp": timestamp,
                "Country": country,
                "Project": project,
                "Repository": repo,
                "Source Branch": source,
                "Target Branch": target,
                "Status": "SYNCED",
                "Commit ID": "N/A",
                "Description": "All changes merged",
                "Author": "N/A",
                "Bitbucket Link": "N/A"
            })
            
        # Add to HTML
        html += f"""
            <tr>
                <td>{country} ({project})</td>
                <td>{repo}</td>
                <td>{source}</td>
                <td>{target}</td>
                <td>{status_html}</td>
                <td>{details_html}</td>
            </tr>
        """
        
    html += """
        </table>
    </body>
    </html>
    """
    
    # Save HTML
    report_path_html = f"audit_report_{timestamp_file}.html"
    with open(report_path_html, "w") as f:
        f.write(html)
    print(f"[+] HTML Report generated: {os.path.abspath(report_path_html)}")
    
    # -------------------------------------------------------------
    # 2. EXCEL REPORT
    # -------------------------------------------------------------
    if excel_rows:
        df = pd.DataFrame(excel_rows)
        report_path_xlsx = f"audit_report_{timestamp_file}.xlsx"
        
        # Create Excel Writer
        writer = pd.ExcelWriter(report_path_xlsx, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Audit Results')
        
        workbook = writer.book
        worksheet = writer.sheets['Audit Results']
        
        # Apply Hyperlinks to the "Bitbucket Link" column
        # Column 'K' (11th column) is Bitbucket Link
        link_col_idx = df.columns.get_loc("Bitbucket Link") + 1 
        
        for i, row in enumerate(df.itertuples(), start=2): # 1-based index, +header
            link = row._asdict()["Bitbucket Link"] # Access by name safely
            val = row._asdict()["Commit ID"]
            
            if link != "N/A":
                # Make the Commit ID a hyperlink
                cell = worksheet.cell(row=i, column=df.columns.get_loc("Commit ID")+1)
                cell.hyperlink = link
                cell.value = val
                cell.style = "Hyperlink"
        
        # Auto-adjust columns width
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        writer.close()
        print(f"[+] Excel Report generated: {os.path.abspath(report_path_xlsx)}")
    else:
        print("[-] No data to write to Excel.")

def main():
    # 1. Load Config
    try:
        # Load credentials from common/config.ini
        common_config = load_config('../common/config.ini')
        bb_url = common_config.get('bitbucket', 'url').rstrip('/')
        bb_user = common_config.get('bitbucket', 'user')
        bb_token = common_config.get('bitbucket', 'token')
        
        # Load Audit Rules
        with open('audit_config.json', 'r') as f:
            audit_rules = json.load(f)
            
    except Exception as e:
        print(f"[-] Config Failure: {e}")
        return

    print("=== Starting Bitbucket Branch Audit (Excel + HTML) ===")
    results = []
    
    for rule in audit_rules:
        project = rule['project']
        repo = rule['repository']
        source = rule['source_branch']
        country = rule.get('country', 'Unknown')
        
        for target in rule['target_branches']:
            print(f"[*] Checking {repo}: {source} -> {target}...")
            
            # Simulated check if URL is dummy (for testing without real Bitbucket)
            if "bitbucket.org" in bb_url and "dummy" in bb_token:
                # Mock Data for demonstration
                missing_commits = [
                    {"id": "abc123456789", "displayId": "abc1234", "message": "Fix NullPointerException in Payment", "author": {"name": "Dev A"}},
                    {"id": "def567890123", "displayId": "def5678", "message": "Update security lib version", "author": {"name": "Dev B"}}
                ] if "branch-B" in target else [] # Simulate B has missing, C is clean
            else:
                missing_commits = check_branch_delta(bb_url, bb_user, bb_token, project, repo, source, target)
            
            check_result = {
                'country': country,
                'project': project,
                'repo': repo,
                'source': source,
                'target': target,
                'deltas': missing_commits if missing_commits else []
            }
            results.append(check_result)
            
    generate_report(results, base_url=bb_url)

if __name__ == "__main__":
    main()
