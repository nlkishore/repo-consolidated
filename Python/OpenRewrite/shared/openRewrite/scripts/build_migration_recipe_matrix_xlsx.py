"""
Build OpenRewrite migration matrix Excel for senior management / estimation review.
Output: C:\\openRewrite\\OpenRewrite-Migration-Recipe-Matrix.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "OpenRewrite-Migration-Recipe-Matrix.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
SUBHEADER_FONT = Font(bold=True)
THIN = Side(style="thin", color="A6A6A6")


def style_header_row(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_subheader_row(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def add_table(ws, start_row: int, headers: list, rows: list) -> int:
    ncol = len(headers)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, ncol)
    r = start_row + 1
    for row in rows:
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        r += 1
    return r


def autosize(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            if cell.value is None:
                continue
            maxlen = max(maxlen, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 12), max_width)


def main():
    wb = Workbook()

    # --- Sheet: Executive_Summary ---
    ws0 = wb.active
    ws0.title = "Executive_Summary"
    ws0["A1"] = "OpenRewrite migration matrix — summary for review"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"] = (
        "Source references: C:\\openRewrite (rewrite.yml, *-CONSTRAINTS.md, rewrite-*.yml). "
        "Automation = recipe can apply changes; Partial = review/tests required; "
        "Manual = backlog / not covered by recipes in this workspace."
    )
    ws0["A2"].alignment = Alignment(wrap_text=True)

    exec_headers = [
        "Category",
        "Migration scope",
        "Primary automated recipes (workspace / upstream)",
        "Automation coverage (high-level)",
        "Manual / backlog (high-level)",
        "Typical risk",
        "Key reference files under C:\\openRewrite",
    ]
    exec_rows = [
        (
            "Java EE → Jakarta EE",
            "javax.* → jakarta.* (selected enterprise APIs)",
            "com.uob.openrewrite.JavaxToJakartaNamespaces; org.openrewrite.java.migrate.jakarta.JavaxMigrationToJakarta",
            "High for Java sources; partial for XML/descriptors",
            "Non-Java resources, third-party JARs still on javax.*, server config",
            "Medium — compile + integration test",
            "rewrite.yml; OPENREWRITE-LEGACY-TO-JAKARTAEE-JDK17.md",
        ),
        (
            "Torque 3.x → Torque 7.x",
            "ORM + WAR integration",
            "com.uob.openrewrite.Torque3To7WithJakartaWebTier; Torque3To7OptionalNewCriteriaPackage (opt-in risky); Torque3To7ManualBacklog",
            "Low–medium (Jakarta web tier + optional Criteria rename); generator/schema not automated",
            "Village removal, peer regeneration, schema XSD, semantic Criteria, build coordinates",
            "High without regenerated peers and DB tests",
            "rewrite-torque3-to-7-webapp.yml; OPENREWRITE-TORQUE3-TO-7-RECIPE-AND-CONSTRAINTS.md; TORQUE-3-TO-7-UPGRADE-SUMMARY.md",
        ),
        (
            "Turbine 3.x / 4.x → Turbine 7.x",
            "Portal servlet pipeline + Fulcrum",
            "com.uob.openrewrite.Turbine4To7WebCompatible; Turbine4To7ManualBacklog; (template) com.misys.rewrite.Turbine3To7PackageUpgrade",
            "Medium for deterministic Java/Jakarta; valves/bootstrap not fully automated",
            "AbstractValve→Valve, Turbine subclass, pipeline XML, Log4j2, dependencies, JSP",
            "High for custom valves and framework forks",
            "rewrite-turbine4-to-7-webapp.yml; OPENREWRITE-TURBINE4-TO-7-RECIPE-AND-CONSTRAINTS.md; misys-openrewrite-template\\rewrite\\",
        ),
        (
            "Spring Framework 4.x → 7.x",
            "Spring 6 (Jakarta) then Spring 7",
            "Phase A: com.uob.openrewrite.SpringFramework4xTo6xJakarta (+ optional WithSpringSecurity6). Phase B (upstream): org.openrewrite.java.spring.framework.UpgradeSpringFramework_7_0",
            "High for many Spring API + POM changes; partial for XML/Security/OAuth",
            "JDK 17+, Hibernate 3/4 removal, JSP/JSTL, XML config, third-party starters, Spring Security semantics",
            "High — multi-phase program",
            "rewrite-spring-4-to-6-jakarta.yml; OPENREWRITE-SPRING4-TO-6-RECIPE-AND-CONSTRAINTS.md",
        ),
        (
            "Cross-cutting",
            "JDK / build / CI",
            "rewrite-migrate-java (Java version recipes — see LEGACY doc); not all wired in this pom",
            "Varies by recipe",
            "Toolchain, Docker, vendor libs, performance/regression",
            "Medium",
            "OPENREWRITE-LEGACY-TO-JAKARTAEE-JDK17.md; pom.xml (plugin + recipe classpath)",
        ),
    ]
    r = add_table(ws0, 4, exec_headers, exec_rows)
    ws0.cell(row=r + 1, column=1, value="Estimation columns for management:")
    ws0.cell(row=r + 2, column=1, value="Add portfolio-specific columns (e.g. FTE weeks, cost, priority) to detail sheets or duplicate this sheet.")
    autosize(ws0)

    def detail_sheet(
        title: str,
        intro: str,
        auto_headers: list,
        auto_rows: list,
        manual_headers: list,
        manual_rows: list,
    ):
        ws = wb.create_sheet(title[:31])  # Excel sheet name limit
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = intro
        ws["A2"].alignment = Alignment(wrap_text=True)
        row = 4
        ws.cell(row=row, column=1, value="SECTION A — Automated / semi-automated recipes")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1
        row = add_table(ws, row, auto_headers, auto_rows)
        row += 2
        ws.cell(row=row, column=1, value="SECTION B — Manual backlog (not fully automatable)")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1
        row = add_table(ws, row, manual_headers, manual_rows)
        row += 2
        ws.cell(
            row=row,
            column=1,
            value="SECTION C — Estimation / sign-off (fill for senior management review)",
        )
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1
        est_headers = [
            "Backlog item / work package",
            "Owner role",
            "Estimate (days)",
            "Confidence",
            "Dependencies",
            "Status",
            "Reviewer sign-off",
        ]
        for j, h in enumerate(est_headers, start=1):
            ws.cell(row=row, column=j, value=h)
        style_header_row(ws, row, len(est_headers))
        for i in range(10):
            rr = row + 1 + i
            for j in range(1, len(est_headers) + 1):
                ws.cell(row=rr, column=j, value="")
                ws.cell(row=rr, column=j).border = Border(
                    left=THIN, right=THIN, top=THIN, bottom=THIN
                )
        autosize(ws)

    # --- Java EE → Jakarta ---
    detail_sheet(
        "JavaEE_to_JakartaEE",
        "Recipes from C:\\openRewrite\\rewrite.yml — javax → jakarta for common enterprise APIs in Java sources.",
        [
            "Recipe display name",
            "Recipe FQN",
            "Automation level",
            "Applies to",
            "Notes / upstream",
        ],
        [
            (
                "Migrate javax namespaces to jakarta (selected enterprise APIs)",
                "com.uob.openrewrite.JavaxToJakartaNamespaces",
                "Full (Java)",
                "Application Java sources in configured modules",
                "Composite: JavaxMigrationToJakarta + explicit ChangePackage rules",
            ),
            (
                "Jakarta migration (broad)",
                "org.openrewrite.java.migrate.jakarta.JavaxMigrationToJakarta",
                "Full (Java, broad)",
                "Same",
                "Included inside JavaxToJakartaNamespaces",
            ),
            (
                "Explicit servlet migration",
                "org.openrewrite.java.ChangePackage javax.servlet → jakarta.servlet",
                "Full",
                "Java",
                "Part of JavaxToJakartaNamespaces",
            ),
            (
                "Explicit JPA migration",
                "org.openrewrite.java.ChangePackage javax.persistence → jakarta.persistence",
                "Full",
                "Java",
                "Part of JavaxToJakartaNamespaces",
            ),
            (
                "Explicit Bean Validation",
                "org.openrewrite.java.ChangePackage javax.validation → jakarta.validation",
                "Full",
                "Java",
                "Part of JavaxToJakartaNamespaces",
            ),
        ],
        [
            "ID",
            "Backlog area",
            "Why manual / partial",
            "Suggested owner",
        ],
        [
            ("J-01", "JSP / taglibs / Facelets", "JSP not covered by default Java-only recipes; taglib URIs must match Jakarta impl", "Web + build"),
            ("J-02", "web.xml / glassfish-web.xml / deployment descriptors", "Schema version, metadata-complete, vendor XML", "App server + integration"),
            ("J-03", "Third-party dependencies", "Libraries still exposing javax.* APIs must be upgraded", "Platform / architecture"),
            ("J-04", "Non-Java config (properties, YAML)", "May need targeted recipes or manual edits", "DevOps / dev"),
            ("J-05", "Generated / vendor sources", "Excluded or overwritten by codegen; align generator first", "Build"),
            ("J-06", "Behavioral API changes", "Some jakarta APIs differ subtly; integration tests required", "QA + dev"),
        ],
    )

    # --- Torque ---
    detail_sheet(
        "Torque3_to_Torque7",
        "Torque ORM 3.x → 7.x per OPENREWRITE-TORQUE3-TO-7-RECIPE-AND-CONSTRAINTS.md and Apache migration guide.",
        [
            "Recipe display name",
            "Recipe FQN",
            "Automation level",
            "Type",
            "Notes",
        ],
        [
            (
                "Torque-era web — Jakarta web tier",
                "com.uob.openrewrite.Torque3To7WithJakartaWebTier",
                "Full (delegates to Jakarta recipe)",
                "Transform",
                "Does not regen Torque peers or remove Village",
            ),
            (
                "OPTIONAL RISKY: util.Criteria → criteria.Criteria",
                "com.uob.openrewrite.Torque3To7OptionalNewCriteriaPackage",
                "Partial / risky",
                "Transform",
                "Semantic differences per Apache docs — opt-in only",
            ),
            (
                "Torque manual backlog detection",
                "com.uob.openrewrite.Torque3To7ManualBacklog",
                "N/A (detection)",
                "Backlog generator",
                "FindTypes + text Find; exportDatatables",
            ),
        ],
        [
            "ID",
            "Backlog area",
            "Description",
            "Suggested owner",
        ],
        [
            ("TQ-01", "Village removal", "Torque 7 runtime does not use Village; redesign data access", "Backend lead"),
            ("TQ-02", "Code generation", "Schema XSD, Maven Torque plugin, templates, regen peers/OM/MapBuilder", "Build + DBA"),
            ("TQ-03", "Criteria semantics", "OR/AND, CUSTOM, doDelete table resolution when moving to criteria.Criteria", "Backend dev"),
            ("TQ-04", "null Connection audit", "Torque 4+ errors on null connection where Torque 3 tolerated", "Backend dev"),
            ("TQ-05", "Torque.properties / datasource XML", "Framework integration (e.g. Turbine component)", "Platform"),
            ("TQ-06", "JDBC / DB scripts", "Dialect-specific scripts if not MySQL-only", "DBA"),
        ],
    )

    # --- Turbine ---
    detail_sheet(
        "Turbine_to_Turbine7",
        "Turbine 3.x/4.x → 7.x: C:\\openRewrite recipes + optional misys-openrewrite-template Turbine3To7 safe set.",
        [
            "Recipe display name",
            "Recipe FQN",
            "Automation level",
            "Location",
            "Notes",
        ],
        [
            (
                "Turbine 4→7 web compatible",
                "com.uob.openrewrite.Turbine4To7WebCompatible",
                "Medium",
                "rewrite.yml",
                "Jakarta + TemplateInfo ChangeType + format",
            ),
            (
                "Turbine manual backlog",
                "com.uob.openrewrite.Turbine4To7ManualBacklog",
                "Detection",
                "rewrite.yml",
                "AbstractValve, Turbine, modules, createRuntimeDirectories Find",
            ),
            (
                "Turbine 3→7 package upgrade (safe subset)",
                "com.misys.rewrite.Turbine3To7PackageUpgrade",
                "Medium",
                "misys-openrewrite-template\\rewrite\\rewrite-turbine7-upgrade.yml",
                "Narrower servlet Jakarta + TemplateInfo; use if aligned with mapping table",
            ),
            (
                "Turbine manual backlog (template)",
                "com.misys.rewrite.Turbine3To7ManualBacklog",
                "Detection",
                "misys-openrewrite-template\\rewrite\\rewrite-turbine7-manual-backlog.yml",
                "Additional FindTypes for services/modules",
            ),
        ],
        [
            "ID",
            "Backlog area",
            "Description",
            "Suggested owner",
        ],
        [
            ("TB-01", "Custom pipeline valves", "AbstractValve removed in Turbine 7; implement Valve + Jakarta types", "Framework dev"),
            ("TB-02", "Turbine servlet subclass", "createRuntimeDirectories vs configureApplication / Path / JAXB", "Framework dev"),
            ("TB-03", "turbine-classic-pipeline.xml", "DefaultSetEncodingValve ordering", "DevOps + dev"),
            ("TB-04", "componentConfiguration / roles", "Fulcrum component class compatibility", "Platform"),
            ("TB-05", "Logging", "Log4j 1.x properties → Log4j2 XML", "Ops"),
            ("TB-06", "Dependencies", "turbine + Fulcrum + Torque versions + Servlet 5 container", "Architecture"),
        ],
    )

    # --- Spring 4 → 7 (two-phase) ---
    detail_sheet(
        "Spring4_to_Spring7",
        "Roadmap: Phase A — Spring 6 + Jakarta (wired in C:\\openRewrite). Phase B — Spring 7 (upstream OpenRewrite; requires compatible rewrite-spring version).",
        [
            "Recipe display name",
            "Recipe FQN",
            "Automation level",
            "Phase",
            "Notes",
        ],
        [
            (
                "Spring → 6.0 + Jakarta namespaces",
                "com.uob.openrewrite.SpringFramework4xTo6xJakarta",
                "High (Java + POM)",
                "A — Spring 6",
                "UpgradeSpringFramework_6_0 + JavaxToJakartaNamespaces",
            ),
            (
                "Spring → 6.0 + Jakarta + Spring Security 6",
                "com.uob.openrewrite.SpringFramework4xTo6xWithSpringSecurity6",
                "High (partial Security)",
                "A — Spring 6",
                "Adds UpgradeSpringSecurity_6_0",
            ),
            (
                "Spring 4→6 manual backlog detection",
                "com.uob.openrewrite.SpringFramework4To6ManualBacklog",
                "Detection",
                "A — Spring 6",
                "Legacy MVC/JDBC/Hibernate3-4, javax in XML/JSP",
            ),
            (
                "Migrate to Spring Framework 7.0 (upstream)",
                "org.openrewrite.java.spring.framework.UpgradeSpringFramework_7_0",
                "High (with caveats)",
                "B — Spring 7",
                "Run after 6.x stable; may require newer rewrite-spring than pinned in pom; includes e.g. JUnit/Jackson upgrades per OpenRewrite docs",
            ),
            (
                "Migrate to Spring Security 7.0 (upstream)",
                "org.openrewrite.java.spring.security7.UpgradeSpringSecurity_7_0",
                "Partial",
                "B — Spring 7",
                "If applicable; coordinate with Spring Framework 7 / Boot 4 stack",
            ),
        ],
        [
            "ID",
            "Backlog area",
            "Description",
            "Suggested owner",
        ],
        [
            ("SP-01", "JDK 17+", "Spring 6+ baseline; CI and runtime images", "Platform"),
            ("SP-02", "Servlet 5+ container", "Tomcat 10+ / Jetty 11+ / EE9+ for WARs", "Ops"),
            ("SP-03", "Hibernate 3/4 Spring support", "Removed in Spring 6; migrate ORM stack", "Architecture"),
            ("SP-04", "SimpleJdbcTemplate / WebMvcConfigurerAdapter", "Removed/deprecated patterns; refactor", "Backend dev"),
            ("SP-05", "Spring XML / jee: namespace", "Bean definitions and JNDI", "Backend dev"),
            ("SP-06", "JSP/JSTL Jakarta", "View layer alignment", "Web dev"),
            ("SP-07", "OAuth2 / SAML / custom security", "Beyond recipe defaults", "Security architect"),
            ("SP-08", "Third-party Spring starters", "Camel, CXF, vendor — Jakarta-compatible versions", "Architecture"),
            ("SP-09", "Spring 7 follow-on", "Validate rewrite-spring version; run UpgradeSpringFramework_7_0 separately", "Tech lead"),
        ],
    )

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
