"""Write openpyxl workbook: AllFiles + Collisions."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from java_component_inventory.scanner import JavaFileRecord


def write_report(
    records: list[JavaFileRecord],
    collisions: dict[str, list[JavaFileRecord]],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet 1: all files (FolderName = component root folder for clarity + parent dir in extra column)
    ws1 = wb.active
    ws1.title = "AllFiles"
    headers = [
        "ComponentFolder",
        "ParentFolderName",
        "PackageName",
        "ClassFileName",
        "FQN",
        "RelativePath",
    ]
    ws1.append(headers)
    for c in ws1[1]:
        c.font = Font(bold=True)
    for r in sorted(records, key=lambda x: (x.component_name, x.relative_path)):
        ws1.append(
            [
                r.component_name,
                r.folder_name,
                r.package_name,
                r.class_file_name,
                r.fqn,
                r.relative_path,
            ]
        )

    # Optional strict 3-column view note in doc / second mini-sheet — user asked 3 columns;
    # we provide ComponentFolder as "folder name under base" which maps to configured component.
    ws3 = wb.create_sheet("Summary3Col")
    ws3.append(["FolderName", "PackageName", "ClassFileName"])
    for c in ws3[1]:
        c.font = Font(bold=True)
    for r in sorted(records, key=lambda x: (x.component_name, x.relative_path)):
        ws3.append([r.component_name, r.package_name, r.class_file_name])

    if collisions:
        ws2 = wb.create_sheet("Collisions")
        ws2.append(["FQN", "OccurrenceCount", "Components", "Paths"])
        for c in ws2[1]:
            c.font = Font(bold=True)
        for fqn, items in sorted(collisions.items(), key=lambda kv: kv[0]):
            comps = "; ".join(sorted({i.component_name for i in items}))
            paths = "; ".join(
                f"{i.component_name}/{i.relative_path}" for i in sorted(items, key=lambda x: x.relative_path)
            )
            ws2.append([fqn, len(items), comps, paths])
    else:
        ws2 = wb.create_sheet("Collisions")
        ws2.append(["FQN", "OccurrenceCount", "Components", "Paths"])
        ws2.append(["(no duplicate FQNs across scanned components)", "", "", ""])

    meta = wb.create_sheet("Meta")
    meta.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    meta.append(["TotalJavaFiles", len(records)])
    meta.append(["CollisionCount", len(collisions)])

    wb.save(output_path)
